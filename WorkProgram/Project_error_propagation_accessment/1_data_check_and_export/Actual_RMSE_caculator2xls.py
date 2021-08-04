import os
import openpyxl
import datetime

# 计算样本实际RMSE的脚本

input_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_error_simulated_formation_energy_on_Ni\raw_txt\Formation_Energy'
# input_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_error_simulated_formation_energy_on_Ni\raw_txt\test'
# input_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_error_simulated_formation_energy_on_Ni\raw_txt\Formation_Energy_500size'
input_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\3_script\0_generate_energy\test'

output_path = './'

size = '2000size'
# size = '250size'


true_energy_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_origin_formation_energy_on_Ni/energies_avg.txt'


def energy_read(catmap_input_file):  # 创建一个species对formation energy的dictionary
    with open(catmap_input_file) as f:
        data = f.read().split("\n")
        title = data[0].split("\t")
        species_name = []
        output_data = {}
        for i in data[1:]:
            _data = i.split("\t")
            # FIXME：这里需要排除气态的能量
            if _data[title.index("site_name")] == "gas": continue
            output_data[_data[title.index("species_name")]] = float(_data[title.index("formation_energy")])
    return output_data

def actual_RMSE_calculator(true_list, errored_list):
    true_list.pop(-5)
    errored_list.pop(-5)

    first = list(map(lambda x: round(x[0]-x[1], 3), zip(true_list, errored_list)))
    mean = sum(first)/len(first)
    # print(mean*mean)  
    second = list(map(lambda num:num*num, first))
    third = sum(second)
    RMSE = (third / (len(true_list)))**(1/2)# 有一个物种用不到
    return RMSE

def excel_write(all_dic, output_path):
    # {RMSE_range:{order:actual_RMSE}}
    xls = openpyxl.Workbook()
    sheet = xls.create_sheet(index=0, title='actual RMSE')

    # 这里没有作排序，但是没有关系，因为作图的时候数据是会自动分割排序的
    sheet.cell(1, 2).value = 'order'
    sheet.cell(1, 3).value = 'RMSE'

    row_count = 2
    for i in all_dic.keys():
        for j in sorted(all_dic[i].keys()):
            sheet.cell(row_count, 1).value = i
            sheet.cell(row_count, 2).value = j
            sheet.cell(row_count, 3).value = all_dic[i][j]
            row_count += 1

    file_name = output_path + '/' + 'actual_RMSE_data' + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print('The output path is ' + file_name)
    xls.save(file_name)
    print('excel has been created!')
    print('work complete!')


correct_energy_dic = energy_read(true_energy_path)


# 制作一个{RMSE_range:{order:actual_RMSE}}
# TODO 干脆这个脚本直接能算bias，variance
final = {}
for i in os.listdir(input_path):
    if 'output_RMSE' in i:
        print('processing ' + i)
        parts = i.split('_')
        RMSE_tag = parts[2] + '_' + parts[3]
        order_RMSE = {}

        for j in os.listdir(input_path + '/' + i):
            true_energy_list = []
            errored_energy_list = []
            order = eval(j.replace('energies', '').replace('.txt', ''))
            errored_energy_dic = energy_read(input_path + '/' + i + '/' + j)
            for species in correct_energy_dic.keys():
                true_energy_list.append(correct_energy_dic[species])
                errored_energy_list.append(errored_energy_dic[species])
            RMSE = actual_RMSE_calculator(true_energy_list, errored_energy_list)
            order_RMSE[order] = RMSE

        final[RMSE_tag] = order_RMSE
# exit(0)
excel_write(final, output_path)
#final字典的格式：{RMSE_range:{order:actual_RMSE}}
