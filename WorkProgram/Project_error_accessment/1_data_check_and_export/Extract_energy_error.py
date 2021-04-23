import os
import pickle
import openpyxl
import datetime
import math
import copy
import pandas as pd
# 读取并输出偏差值，写成excel，方便处理，也可以单独取出
# 写成excel，预计是一个含多个sheet（由物种名称）的excel或者直接多个excel
# 也可以改成直接输出能量值
# 中间过程中使用了保留三位小数，视情况更改


# Error_ERaw = energy_read('./energies_error_0.1_0.12_0.txt')

# file_path = './errored_energy'
# 改目标errored energy的文件夹们所在文件夹的位置
true_energy = './energies_true.txt'
mode = 'abs'
division = True
decimal_point = 4# round位数
# division_resolution = 0.04# 注意因为round3，因此精度只到0.00
division_resolution = 0.05
division_range = (0, 0.9)
file_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_error_simulated_formation_energy_on_Ni\raw_txt\2000sizeN'
# 为了csv的设置
columns = ['RMSE 0.10-0.12eV',
           'RMSE 0.12-0.14eV',
           'RMSE 0.14-0.16eV',
           'RMSE 0.16-0.18eV',
           'RMSE 0.18-0.20eV',
           'RMSE 0.20-0.22eV',
           'RMSE 0.22-0.24eV',
           'RMSE 0.24-0.26eV',
           'RMSE 0.26-0.28eV',
           'RMSE 0.28-0.30eV',
           ]
title = {'(0.1, 0.12)':'RMSE 0.10-0.12eV', 
        '(0.12, 0.14)':'RMSE 0.12-0.14eV', 
        '(0.14, 0.16)':'RMSE 0.14-0.16eV', 
        '(0.16, 0.18)':'RMSE 0.16-0.18eV', 
         '(0.18, 0.2)':'RMSE 0.18-0.20eV', 
         '(0.2, 0.22)':'RMSE 0.20-0.22eV', 
        '(0.22, 0.24)':'RMSE 0.22-0.24eV', 
        '(0.24, 0.26)':'RMSE 0.24-0.26eV', 
        '(0.26, 0.28)':'RMSE 0.26-0.28eV', 
         '(0.28, 0.3)':'RMSE 0.28-0.30eV',
           }

def excel_write(all_dic, output_path, mode, division):
    '''
    每个sheet由物种名称命名
    sheet中第一行填写RMSE范围，origin应该可以把第一行非数据的名称一并导入; 其他行为error or energy
    也即某张sheet的某一列代表某个物种在某个RMSE范围中的能量误差，数据个数与生成样本size相等

    '''
    if division:
        # all_dic = {'species_name':{'RMSE_range':{divsion_range:count}}}
        RMSE_order = {}
        order = 2
        for key1 in sorted(all_dic.keys()):
            for key2 in sorted(all_dic[key1].keys()):
                RMSE_order[order] = key2
                order += 1

        xls = openpyxl.Workbook()

        index_number = 0
        for i in sorted(all_dic.keys()):
            sheet = xls.create_sheet(index=index_number, title=i)
            index_number += 1

            point_row = 2

            for point in division_point:
                sheet.cell(point_row,1).value = point
                for j in range(len(all_dic[i].keys())):
                    sheet.cell(1, j + 2).value = 'RMSE ' + str(eval(RMSE_order[j + 2])[0]) + '-' + str(eval(RMSE_order[j + 2])[1]) + 'eV' # 写第一行RMSE范围
                    sheet.cell(point_row, j + 2).value = all_dic[i][RMSE_order[j + 2]][point]
                point_row += 1
    else:
        # all_dic = {'species_name':{'RMSE_range':[error or energy]}}
        xls = openpyxl.Workbook()
        index_number = 0
        for i in sorted(all_dic.keys()):
            sheet = xls.create_sheet(index=index_number, title=i)
            index_number += 1
            column_number = 1
            for j in sorted(all_dic[i].keys()):
                sheet.cell(1,column_number).value = 'RMSE ' + str(eval(j)[0]) + '-' + str(eval(j)[1]) + 'eV'# 写第一行RMSE范围
                row_number = 2
                for error in all_dic[i][j]:
                    sheet.cell(row_number, column_number).value = error
                    row_number += 1
                column_number += 1


    file_name = output_path + '/' + 'Error_data_' + mode + '_division' + str(division) + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print('The output path is ' + file_name)
    xls.save(file_name)
    print('excel has been created!')


def csv_write(all_dic, output_path, mode, division, division_point, columns):
    try:
        os.mkdir(output_path + '/' + 'csvs')
    except:
        pass
    output_path = output_path + '/' + 'csvs'


    # all_dic = {'species_name':{'RMSE_range':{divsion_range:count}}}
    if division:
        for species in all_dic:
            content = all_dic[species]
            new_dic = {}
            for RMSE in content:
                new_dic[title[RMSE]] = content[RMSE]
            dataframe = pd.DataFrame(new_dic)
            dataframe.index.name = ('Error distribution of species %s' % species)
            file_name = output_path + '/' + species + '.csv'
            print('The output path is ' + file_name)
            dataframe.to_csv(file_name,index = division_point,sep = ',', columns = columns)
        print('all csv has been created!')
    else:
        print('Function not build')
    



def energy_read(catmap_input_file):  # 创建一个species对formation energy的dictionary
    with open(catmap_input_file) as f:
        data = f.read().split("\n")
        title = data[0].split("\t")
        species_name = []
        output_data = {}
        for i in data[1:]:
            _data = i.split("\t")
            # FIXME：这里需要排除气态的能量
            if _data[title.index("site_name")] == "gas": continue
            output_data[_data[title.index("species_name")]] = float(_data[title.index("formation_energy")])
    return output_data

true_ERaw = energy_read(true_energy)# 正确能量值，这个一般不用变化
error_dic = {}# 整个RMSE误差范围内，物种能量的偏差: {'species_name':[error or energy]}
error_dic_divided = {}# 按RMSE分割范围导出物种能量偏差: {'species_name':{'RMSE_range':[error or energy]}}
for i in true_ERaw.keys():
    error_dic_divided[i] = dict()

for RMSE_group in os.listdir(file_path):
    if 'output_RMSE_' in RMSE_group:
        print(RMSE_group)
        RMSE_range_size = RMSE_group.replace('output_RMSE_', '')
        RMSE_range_size = RMSE_range_size.split('_')
        RMSE_range_size = str((eval(RMSE_range_size[0]), eval(RMSE_range_size[1])))
        for txt in os.listdir(file_path + '/' + RMSE_group):
            Error_ERaw = energy_read(file_path + '/' + RMSE_group + '/' + txt)
            for i in true_ERaw.keys():
                if mode == 'all':
                    e = Error_ERaw[i] - true_ERaw[i]
                    try:
                        error_dic[i].extend([e])
                    except KeyError:
                        error_dic[i] = [e]
                    try:
                        error_dic_divided[i][RMSE_range_size].extend([e])
                    except KeyError:
                        error_dic_divided[i][RMSE_range_size] = [e]
                if mode == 'abs':
                    e = abs(Error_ERaw[i] - true_ERaw[i])
                    try:
                        error_dic[i].extend([e])
                    except KeyError:
                        error_dic[i] = [e]
                    try:
                        error_dic_divided[i][RMSE_range_size].extend([e])
                    except KeyError:
                        error_dic_divided[i][RMSE_range_size] = [e]
                if mode == 'positive':
                    e = Error_ERaw[i] - true_ERaw[i]
                    if e > 0:
                        try:
                            error_dic[i].extend([e])
                        except KeyError:
                            error_dic[i] = [e]
                        try:
                            error_dic_divided[i][RMSE_range_size].extend([e])
                        except KeyError:
                            error_dic_divided[i][RMSE_range_size] = [e]
                    else:
                        continue
                if mode == 'energy':
                    e = Error_ERaw[i]
                    try:
                        error_dic[i].extend([e])
                    except KeyError:
                        error_dic[i] = [e]
                    try:
                        error_dic_divided[i][RMSE_range_size].extend([e])
                    except KeyError:
                        error_dic_divided[i][RMSE_range_size] = [e]
# print(len(error_dic_divided['OH']['0.16_0.18']))
# exit(0)
if division:
    '''
    需要division的话，就要做成一个字典{'species':{'RMSE_range':{0:0, division_range_low:count}}}
    然后写成excel，excel和原来的形式没有太大区别，每个sheet是物种，第一列是分割范围到时候作为orginlab中的x轴，其他每个范围的count作列，届时作y轴
    '''
    print('')
    print('')
    print('Division carrying out')
    print('')
    # 先制作目标dic的格式
    error_dic_divided_division = copy.deepcopy(error_dic_divided)    

    division_range_min = division_range[0]
    division_range_max = division_range[1]
    interval_num = math.ceil((division_range_max - division_range_min)/division_resolution)
    division_interval = []
    counter = 0
    count_dic = {}# 初始化count的dic
    division_point = []

    for i in range(interval_num):# 制作
        division_range_low = division_range_min + counter*division_resolution
        division_range_high = division_range_min + (counter + 1)*division_resolution
        counter += 1
        # division_interval.append((division_range_low, division_range_high))
        division_interval.append((round(division_range_low, decimal_point), round(division_range_high, decimal_point)))# 这次直接用round抵消掉python计算的问题，有需求的话再变化
        # count_dic[division_range_low, 3] = 0
        count_dic[round((division_range_low + division_range_high)/2, decimal_point)] = 0
        division_point.append(round((division_range_low + division_range_high)/2, decimal_point))

    for key1 in error_dic_divided.keys():# 物种
        for key2 in error_dic_divided[key1].keys():# RMSE范围
            print('Processing ' + key1 + ' in ' + str(key2))
            work = copy.deepcopy(count_dic)
            for e in error_dic_divided[key1][key2]:# 物种在某RMSE范围中的error或者energy
                for r in division_interval:
                    if e < r[1] and e >= r[0]:
                        work[round((r[0] + r[1])/2, 3)] = work[round((r[0] + r[1])/2, decimal_point)] + 1
                error_dic_divided_division[key1][key2] = work


    # csv_write(error_dic_divided_division, file_path, mode, division, division_point, columns)
    excel_write(error_dic_divided_division, file_path, mode, division)                    
else:
    excel_write(error_dic_divided, file_path, mode, division)

print('work complete!')
exit(0)

# 输出C，所有RMSE范围下的能量
# print(error_dic['C'])
for i in error_dic['OH']:
    print(i)

