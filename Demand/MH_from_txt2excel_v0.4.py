import os
import datetime
import openpyxl
import sys
import io
import copy

# v0.2 修复主要主题词和扩展主题词定义错误的问题

# v0.3 增加功能将category/*xxx类主题词识别成主要主题词category/xxx;
#      修复将PMID字符全都识别为一种词条，现在只认为PMID-字符是一种词条;
#      增加PMID词条核查功能，供使用者检查程序是否正确识别PMID个数

# v0.4 增加功能将*category/xxx类主题词识别成主要主题词category/xxx，暂时认为*category/xxx与category/*xxx相同;
#      修复程序会将空行也识别成一个词条

# v0.5 代码简化TODO; 实现批量处理每个txt，生成1个总excel表和每个txt的对应excel表，也即n个txt生产n+1个excelTODO

# 使用前请先阅读注释
# 请使用python3运行此脚本
# 重要！请注意泛用性，该脚本遇到不同格式的txt文件时可能需要根据需求进行一定修改

# 功能：可以批量将某个文件下特定格式txt中的内容，按照要求输入并生成对应excel文件

# 请输入txt文件所在文件夹
files_path = r'C:\Users\OldKuroCat\Desktop\private\try'
# 请输入生成excel文件的路径
output_path = r'C:\Users\OldKuroCat\Desktop\private'

'''
-------------------------------------------------------------------------------------------
'''

def excel_write(all_dic, output_path): # 可能需要加order_list。不需要了，excel可以自动匹配
    # 注意该函数写入excel使用的是openpyxl，而不是xlwt（由于string上限问题）
    xls = openpyxl.Workbook()
    sheet = xls.create_sheet(index=0, title="test")

    # 先建立分类栏：PMID, 主要主题词, 扩展主题词, 所有主题词
    # 使用xlwt的话，坐标分别是(0,0), (0,1), (0,2), (0,3)
    # 使用openpyxl的话，坐标分别+1（从1开始计算）
    sheet.cell(1,1).value = 'PMID'
    sheet.cell(1,2).value = '主要主题词'
    sheet.cell(1,3).value = '扩展主题词'
    sheet.cell(1,4).value = '所有主题词'

    start_row = 2
    for PMID in all_dic:# {PMID:{Mian_MH: , Expand_MH: , All_MH: }}
        sheet.cell(start_row,1).value = PMID
        for MH_type in all_dic[PMID]:
            if MH_type == 'Main_MH':
                sheet.cell(start_row,2).value = all_dic[PMID][MH_type]
            if MH_type == 'Expand_MH':
                sheet.cell(start_row,3).value = all_dic[PMID][MH_type]
            if MH_type == 'All_MH':
                sheet.cell(start_row,4).value = all_dic[PMID][MH_type]
        start_row += 1

    file_name = output_path + '/' + 'MH' + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print('The output path is ' + file_name)
    xls.save(file_name)
    print('excel has been created!')
    print('work complete!')

# TODO 简化代码
# 这个函数不能直接使用，没有达到功能，必须集成进另一个函数里
def dicSave(PMID_Mh_dic, PMID, MH_type, Category, piece):
    try:# 如果还未创建扩展主题词key，先创建，再存value。否则直接用分号连接value
        PMID_Mh_dic[PMID][MH_type] = PMID_Mh_dic[PMID][MH_type] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
    except KeyError:
        PMID_Mh_dic[PMID][MH_type] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
    return PMID_Mh_dic
# def collectMH2dic(tag_content):
#     def dicSave():
#         pass
#     PMID_Mh_dic = {}
#     pass


sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')# 修改标准输出，避免gbk无法编译一些符号

files_list = os.listdir(files_path)# 读取目标文件夹下所有目标文件
# print(files_list)

for txt_name in files_list:
    with open(files_path+'/'+txt_name,'r',encoding = 'utf8') as f:
        content = f.readlines()
    
    # 用来测试文档中有多少个片段，这次是10个，那么all_param的长度应该为10
    at = " ".join(content)
    count_PMID = at.count('PMID-')
    
    # count_PMID = 0
    # for i in content:
    #     if i.startwith('PMID-',0,5)

    # exit(0)

    working_content = copy.deepcopy(content)
    all_param = []
    one_lecture_list = []
    while len(working_content) != 0:# 先将每个片段提取出来，分list，并装在一个大list里。为了避免未来需要更多内容，不直接进行信息提取，只分段。
        current_line = working_content.pop(0)
        if current_line != '\n':
            one_lecture_list.append(current_line)
        else:
            all_param.append(one_lecture_list)
            one_lecture_list = []

    if one_lecture_list != []:# 以防最后一个片段没有换行分割
        all_param.append(one_lecture_list)

    # print('######')
    # print(all_param)
    # print('######')

    # 去掉误识别的空行片段
    all_param = list(filter(None, all_param))

    # 验证处理，以防txt格式变化造成问题
    # TODO 如果出问题，给出已slice的PMID，估计一时半会儿不需要
    passage_number = len(all_param)
    if len(all_param) != count_PMID:
        print('WARNING: slicing may not reliable!')
        print(('program get %s passage, however, there are %s of PMIDs.') % (passage_number, count_PMID))
        print('PLEASE check or contact author!')
    if len(all_param) == count_PMID:
        print(("Result: program get %s of PMIDs' passage!") % passage_number)

    # 提取PMID作key，[[主要主题词],[扩展主题词],[所有主题词]]。注意主要主题词要把*去掉。
    PMID_Mh_dic = {}
    for every_passage in all_param:
        PMID = ''# 以防未来txt格式出现改变，但程序能正常执行，导致无法检查出错误
        for tag_content in every_passage:
            if 'DA' in tag_content:# 有一个不需要的MHDA
                continue
            if 'PMID-' in tag_content:
                PMID = tag_content.replace('-','').strip('PMID').strip()
                PMID_Mh_dic[PMID] = dict()
            if 'MH  -' in tag_content:# 按需求收集主题词，用;连接
                if '/' in tag_content:# 带/的是一种特殊格式，比如A/B/C意味着两个MH，也即是A/B和A/C
                    # print(tag_content)
                    tag_content = tag_content.strip('MH').replace('-','').strip()
                    tag_content_list = tag_content.split('/')
                    Category = tag_content_list.pop(0)
                    if '*' in Category:
                        Category = Category.replace('*','')
                        for piece in tag_content_list:
                            piece = '*' + piece
                            if '*' not in piece:# 不带*的是扩展主题词
                                # dicSave(PMID_Mh_dic, PMID, MH_type, Category, piece)
                                try:# 如果还未创建扩展主题词key，先创建，再存value。否则直接用分号连接value
                                    PMID_Mh_dic[PMID]['Expand_MH'] = PMID_Mh_dic[PMID]['Expand_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except KeyError:
                                    PMID_Mh_dic[PMID]['Expand_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                    
                                try:
                                    PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except:
                                    PMID_Mh_dic[PMID]['All_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                            else:
                                try:# 如果还未创建主要主题词key，先创建，再存value。否则直接用分号连接value
                                    PMID_Mh_dic[PMID]['Main_MH'] = PMID_Mh_dic[PMID]['Main_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except KeyError:
                                    PMID_Mh_dic[PMID]['Main_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                try:
                                    PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except:
                                    PMID_Mh_dic[PMID]['All_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                    else:
                        for piece in tag_content_list:
                            if '*' not in piece:# 不带*的是扩展主题词
                                try:# 如果还未创建扩展主题词key，先创建，再存value。否则直接用分号连接value
                                    PMID_Mh_dic[PMID]['Expand_MH'] = PMID_Mh_dic[PMID]['Expand_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except KeyError:
                                    PMID_Mh_dic[PMID]['Expand_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                    
                                try:
                                    PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except:
                                    PMID_Mh_dic[PMID]['All_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                            else:
                                try:# 如果还未创建主要主题词key，先创建，再存value。否则直接用分号连接value
                                    PMID_Mh_dic[PMID]['Main_MH'] = PMID_Mh_dic[PMID]['Main_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except KeyError:
                                    PMID_Mh_dic[PMID]['Main_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                try:
                                    PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                                except:
                                    PMID_Mh_dic[PMID]['All_MH'] = Category + '/' + piece.strip('MH').replace('-','').strip().replace('*','').strip('\n')

                else:
                    if '*' not in tag_content:# 不带*的是扩展主题词
                        try:# 如果还未创建扩展主题词key，先创建，再存value。否则直接用分号连接value
                            PMID_Mh_dic[PMID]['Expand_MH'] = PMID_Mh_dic[PMID]['Expand_MH'] + '; ' + tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                        except KeyError:
                            PMID_Mh_dic[PMID]['Expand_MH'] = tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                            
                        try:
                            PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                        except:
                            PMID_Mh_dic[PMID]['All_MH'] = tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                    else:
                        try:# 如果还未创建主要主题词key，先创建，再存value。否则直接用分号连接value
                            PMID_Mh_dic[PMID]['Main_MH'] = PMID_Mh_dic[PMID]['Main_MH'] + '; ' + tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                        except KeyError:
                            PMID_Mh_dic[PMID]['Main_MH'] = tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                        try:
                            PMID_Mh_dic[PMID]['All_MH'] = PMID_Mh_dic[PMID]['All_MH'] + '; ' + tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')
                        except:
                            PMID_Mh_dic[PMID]['All_MH'] = tag_content.strip('MH').replace('-','').strip().replace('*','').strip('\n')


    # print(PMID_Mh_dic['31295471'])

excel_write(PMID_Mh_dic, output_path)