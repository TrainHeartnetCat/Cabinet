import os
import sys
import io
import langid
import datetime
import openpyxl
from openpyxl.styles import PatternFill


# version 0.2
# 使用前请先阅读注释
# 请使用python3运行此脚本
# 此脚本的运行相对耗时较长，内存使用较大，请耐心等待
# 由于excel_write函数在创建同名文件时会报错，为了方便使用，excel生成时会自动加上当前时间的后缀
# 重要！由于excel写入单元格的信息长度有限制，因此过长的内容会被截断
# 重要！请注意泛用性，该脚本遇到不同类型html文件时可能需要根据需求进行一定修改


# 功能：可以批量将某个文件下特定格式html中的内容，按照要求输入并生成对应excel文件

# 请输入html文件所在文件夹
files_path = r'C:\Users\OldKuroCat\Desktop\private\需求1\2020-09-16 301-400\claim'
# 请输入生成excel文件的路径
output_path = r'C:\Users\OldKuroCat\Desktop\private\需求1\2020-09-16 301-400'

'''
-------------------------------------------------------------------------------------------
'''




def txt_wrap_by(start_str, end, html):
    start = html.find(start_str)
    if start >= 0:
        start += len(start_str)
        end = html.find(end, start)
        if end >= 0:
            return html[start:end].strip()

def excel_write(all_dic, output_path): # 可能需要加order_list。不需要了，excel可以自动匹配
    # 注意该函数写入excel使用的是openpyxl，而不是xlwt（由于string上限问题）
    xls = openpyxl.Workbook()
    sheet = xls.create_sheet(index=0, title="test")

    # 先建立分类栏：公开（公告）号，权利要求-中文，权利要求-英语，权利要求-其他语言
    # 使用xlwt的话，坐标分别是(0,0), (0,1), (0,2), (0,3)
    # 使用openpyxl的话，坐标分别+1（从1开始计算）
    sheet.cell(1,1).value = '公开（公告）号'
    sheet.cell(1,2).value = '权利要求-中文'
    sheet.cell(1,3).value = '权利要求-英语'
    sheet.cell(1,4).value = '权利要求-其他语言'
    # sheet.write(0,0, label = '公开（公告）号')
    # sheet.write(0,1, label = '权利要求-中文')
    # sheet.write(0,2, label = '权利要求-英语')
    # sheet.write(0,3, label = '权利要求-其他语言')

    # 将公开号随机写入，然后根据语言标签识别内容语言，按分类写入内容
    # all_dic的格式为{'order':[('content', 'language'),...]}
    start_row = 2
    for order in all_dic:
        sheet.cell(start_row,1).value = order
        for item in all_dic[order]:
            if 'CodeL&H: ' in item[0]:
                modified_str = item[0].replace('CodeL&H: ','')
                # print(len(modified_str))
                fille = PatternFill('solid', fgColor = 'FFD700') # 加入黄色填充以示由于长度超出，对字符串进行了阉割处理
                if item[1] == 'zh':
                    sheet.cell(start_row,2).value = 'Too Long'
                    sheet.cell(start_row,2).fill = fille
                elif item[1] == 'en':
                    sheet.cell(start_row,3).value = 'Too Long'
                    sheet.cell(start_row,3).fill = fille
                else:
                    sheet.cell(start_row,4).value = 'Too Long'
                    sheet.cell(start_row,4).fill = fille
            else:
                if item[1] == 'zh':
                    sheet.cell(start_row,2).value = item[0]
                elif item[1] == 'en':
                    sheet.cell(start_row,3).value = item[0]
                else:
                    sheet.cell(start_row,4).value = item[0]
        start_row += 1
    # start_row = 1
    # for order in all_dic:
    #     print(order)
    #     sheet.write(start_row,0, label = order)
    #     for item in all_dic[order]:
    #         if item[1] == 'zh':
    #             sheet.write(start_row,1, label = item[0])
    #         elif item[1] == 'en':
    #             sheet.write(start_row,2, label = item[0])
    #         # else:
    #         #     sheet.write(start_row,3, label = item[0])
    #     start_row += 1
    file_name = output_path + '/' + 'claim' + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print(file_name)
    xls.save(file_name)
    print('excel has been created!')
    print('work complete!')






sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')# 修改标准输出，避免gbk无法编译一些html中的符号
# print('-------------------------------------------------------------------------tag-------------------------------------------------------------------------')
files_list = os.listdir(files_path)# 读取目标文件夹下所有目标文件
# print(files_list)
all_dic = {}
for html_name in files_list:
    print(html_name)
    work_list = []
    with open(files_path+'/'+html_name, 'rb') as f:
        html = f.readlines()
    html.pop(0)# 这批html中第一行不需要，一般来说html的第一行都只是属性，不是正文内容，但注意泛用性
    # print(html_name)
    for message in html:
        str_ele = str(message,encoding='utf8')
        # str_ele = str_ele.replace('<sub>','').replace('</sub>','')
        while txt_wrap_by("<",">",str_ele) != None:
            target = '<' + txt_wrap_by("<",">",str_ele) + '>'
            str_ele = str_ele.replace(target,'')
            # print(len(str_ele))
        if str_ele == '' or str_ele == '\n':
            continue
        else:
            str_ele = str_ele.rstrip('\n')
            str_ele = str_ele.strip()
            # 超出长度剪掉，并做标记指示。经过不完全测试，长度8000以下的字符串可以写入，8500以上不行
            if len(str_ele) > 32767:
                str_ele = str_ele.replace('  ',' ')
                if len(str_ele) > 32767:
                    str_ele = 'CodeL&H: ' + str_ele
                    # print(len(str_ele))
                    # str_ele = str_ele[:8259]
                    # print(len(str_ele))
            work_list.append((str_ele,langid.classify(str_ele)[0]))# 单个html中的一条内容，循环完后获得整个内容
        
    all_dic[os.path.splitext(html_name)[0]] = work_list# 这批html的所有内容集合

excel_write(all_dic, output_path)


# print(all_dic['HRP20191814T1'])


# a = str(html[2],encoding='utf8')
# b' \n'
# regex = r'<([\s\S]*)>'
# matches = re.findall(regex, a)

# for match in matches:
#     # print(match)
#     break
#decode('utf-8','ignore'),encode('utf-8','ignore') 加一个'ignore'参数 就会忽略错误。
