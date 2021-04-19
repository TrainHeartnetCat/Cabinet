import os
import pickle
import openpyxl
import datetime

# 读取并输出偏差值，写成excel，方便处理，也可以单独取出
# 写成excel，预计是一个含多个sheet（由物种名称）的excel或者直接多个excel
# 也可以改成直接输出能量值
def excel_write(all_dic, output_path, mode):# all_dic = {'species_name':{'RMSE_range':[error or energy]}}
    '''
    每个sheet由物种名称命名
    sheet中第一行填写RMSE范围，origin应该可以把第一行非数据的名称一并导入; 其他行为error or energy
    也即某张sheet的某一列代表某个物种在某个RMSE范围中的能量误差，数据个数与生成样本size相等

    '''
    xls = openpyxl.Workbook()
    index_number = 0
    for i in sorted(all_dic.keys()):# TODO 这里的排序是对字符串排序，因此0.1_0.12反而会被扔到后面去。我记得自己以前写过可以fix这个问题的排序方法，以后再回顾一下吧。
        sheet = xls.create_sheet(index=index_number, title=i)
        index_number += 1
        column_number = 1
        for j in sorted(all_dic[i].keys()):
            sheet.cell(1,column_number).value = j# 写第一行RMSE范围
            row_number = 2
            for error in all_dic[i][j]:
                sheet.cell(row_number, column_number).value = error
                row_number += 1
            column_number += 1


    file_name = output_path + '/' + 'Error_data_' + mode + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print('The output path is ' + file_name)
    xls.save(file_name)
    print('excel has been created!')
    print('work complete!')



def energy_read(catmap_input_file):  # 创建一个species对formation energy的dictionary
    with open(catmap_input_file) as f:
        data = f.read().split("\n")
        title = data[0].split("\t")
        species_name = []
        output_data = {}
        for i in data[1:]:
            _data = i.split("\t")
            # FIXME：这里需要排除气态的能量
            if _data[title.index("site_name")] == "gas": continue
            output_data[_data[title.index("species_name")]] = float(_data[title.index("formation_energy")])
    return output_data

true_ERaw = energy_read('./energies_true.txt')# 正确能量值，这个一般不用变化
# Error_ERaw = energy_read('./energies_error_0.1_0.12_0.txt')

# file_path = './errored_energy'
# 改目标errored energy的文件夹们所在文件夹的位置
mode = 'positive'
file_path = r'D:\Works\backup\Project01_ErrorEstimate_DRM_on_Ni111\0_raw_data\All_error_simulated_formation_energy_on_Ni\raw_txt\Formation_Energy'

error_dic = {}# 整个RMSE误差范围内，物种能量的偏差: {'species_name':[error or energy]}
error_dic_divided = {}# 按RMSE分割范围导出物种能量偏差: {'species_name':{'RMSE_range':[error or energy]}}
for i in true_ERaw.keys():
    error_dic_divided[i] = dict()

for RMSE_group in os.listdir(file_path):
    if 'output_RMSE_' in RMSE_group:
        print(RMSE_group)
        RMSE_range_size = RMSE_group.replace('output_RMSE_', '')
        for txt in os.listdir(file_path + '/' + RMSE_group):
            Error_ERaw = energy_read(file_path + '/' + RMSE_group + '/' + txt)
            for i in true_ERaw.keys():
                if mode == 'all':
                    e = Error_ERaw[i] - true_ERaw[i]
                    try:
                        error_dic[i].extend([e])
                    except KeyError:
                        error_dic[i] = [e]
                    try:
                        error_dic_divided[i][RMSE_range_size].extend([e])
                    except KeyError:
                        error_dic_divided[i][RMSE_range_size] = [e]
                if mode == 'abs':
                    e = abs(Error_ERaw[i] - true_ERaw[i])
                    try:
                        error_dic[i].extend([e])
                    except KeyError:
                        error_dic[i] = [e]
                    try:
                        error_dic_divided[i][RMSE_range_size].extend([e])
                    except KeyError:
                        error_dic_divided[i][RMSE_range_size] = [e]
                if mode == 'positive':
                    e = Error_ERaw[i] - true_ERaw[i]
                    if e > 0:
                        try:
                            error_dic[i].extend([e])
                        except KeyError:
                            error_dic[i] = [e]
                        try:
                            error_dic_divided[i][RMSE_range_size].extend([e])
                        except KeyError:
                            error_dic_divided[i][RMSE_range_size] = [e]
                    else:
                        continue
# print(error_dic_divided['0.12_0.14_1000Size']['C'])

excel_write(error_dic_divided, file_path, mode)
exit(0)

# 输出C，所有RMSE范围下的能量
# print(error_dic['C'])
for i in error_dic['OH']:
    print(i)

