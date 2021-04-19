# coding: utf-8
import os
import copy
from sklearn.metrics import mean_squared_error
import numpy as np
import tqdm
import shutil
import matplotlib.pyplot as plt
import random
import openpyxl
import datetime

'''
初始正确能量file中含有一个DRM不涉及的物种H-OH
新使用的energy_true.txt中已经直接剔除该物种
'''
'''
要注意选取的bias是泛化误差的，实际样本的bias是和选取的不完全等的
而生成的RMSE，是样本的RMSE，不是泛化的，注意了！
'''
'''
将原来的保留3位小数，修改为保留7位小数
'''
'''
固定一个随机种子，使得其可重复
'''

# # 3000N
# np.random.seed(963)
# random.seed(5623)

# 2500N
# np.random.seed(16)
# random.seed(266)

# 2000N
np.random.seed(46)
random.seed(101)

# 1500N
# np.random.seed(32)
# random.seed(199)

# 1000N
# np.random.seed(90)
# random.seed(766)

# 500N
# np.random.seed(11)
# random.seed(321)

# 100N
# np.random.seed(2)
# random.seed(67)

# 50N
# np.random.seed(6698)
# random.seed(77555)

def energy_read(catmap_input_file):  # 创建一个species对formation energy的dictionary
    with open(catmap_input_file) as f:
        data = f.read().split("\n")
        title = data[0].split("\t")
        species_name = []
        output_data = {}
        for i in data[1:]:
            _data = i.split("\t")
            # FIXME：这里需要排除气态的能量
            if _data[title.index("site_name")] == "gas": continue
            output_data[_data[title.index("species_name")]] = float(_data[title.index("formation_energy")])
    return output_data

def excel_write(all_dic, output_path):
    # {RMSE_range:{order:(genal_bias, genal_var, actual_bias, actual_var, actual_RMSE, error)}}
    xls = openpyxl.Workbook()
    sheet = xls.create_sheet(index=0, title='actual RMSE')

    # 这里没有作排序，但是没有关系，因为作图的时候数据是会自动分割排序的
    sheet.cell(1, 1).value = 'RMSE_range'
    sheet.cell(1, 2).value = 'order'
    sheet.cell(1, 3).value = 'genal_bias'
    sheet.cell(1, 4).value = 'genal_variance'
    sheet.cell(1, 5).value = 'sampling_bias'
    sheet.cell(1, 6).value = 'sampling_variance'
    sheet.cell(1, 7).value = 'sampling_RMSE'
    sheet.cell(1, 8).value = 'error'

    row_count = 2
    for i in sorted(all_dic.keys()):# RMSE_range
        for j in sorted(all_dic[i].keys()):
            sheet.cell(row_count, 1).value = str(i)
            sheet.cell(row_count, 2).value = j
            sheet.cell(row_count, 3).value = all_dic[i][j][0]
            sheet.cell(row_count, 4).value = all_dic[i][j][1]
            sheet.cell(row_count, 5).value = all_dic[i][j][2]
            sheet.cell(row_count, 6).value = all_dic[i][j][3]
            sheet.cell(row_count, 7).value = all_dic[i][j][4]
            sheet.cell(row_count, 8).value = str(all_dic[i][j][5])            
            row_count += 1

    file_name = output_path + '/' + 'raw_error_' + str(energy_num) + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    print('The output path is ' + file_name)
    xls.save(file_name)
    print('excel has been created!')
    print('work complete!')

def generate_Random_Number(RMSE_low, RMSE_high, size):
    '''
    根据
    https://www.jianshu.com/p/8c7f033be58a
    中的公式
    E((y-f(x))^2) = Var(f(x)) + Bias(f(x)^2)
    或者是
    https://en.wikipedia.org/wiki/Bias%E2%80%93variance_tradeoff

    最终公式：
    MSE(x) = var(x) + bias(x)^2 + ϵ2 (忽略)
    也就是说
    为了生成某个RMSE的样本，需要先得到MSE
    然后在均匀分布-MSE到+MSE区间范围内给出bias的平方
    减去得到var，也就是方差，标准差的平方

    '''
    while 1:
        # 1. 生成size个数目的均匀误差，每个在一定区间内
        # 这个区间是正负RMSE！
        RMSE = (RMSE_high + RMSE_low) / 2

        MSE = np.square(RMSE)
        bias_square = random.uniform(0,MSE)
        var = MSE - bias_square
        std = np.sqrt(var)
        # 注意bias是有正负的，这里随机选一个正负
        if random.choice([0,1]) == 0:

            bias = np.sqrt(bias_square)
        else:
            bias = - np.sqrt(bias_square)

        random_error = np.random.normal(bias,std,size)

        # 2. 计算size个样本的RMSE，注意不是单个的
        actual_RMSE = np.sqrt(np.mean(np.square(random_error)))

        if RMSE_low < actual_RMSE < RMSE_high:
            # print(random_error)
            # print(actual_RMSE)
            actual_bias = np.mean(random_error)
            actual_variance = np.var(random_error)
            return random_error, (bias, var, actual_bias, actual_variance, actual_RMSE, list(random_error))





excel_output = './ntest/'
# ---Step0 设置读取的能量文件以及输出的能量文件夹，以及生成的误差的方差范围
energy_num = 2000
# energy_num = 1000
RMSE_group = [(0.1, 0.12), (0.12, 0.14), (0.14, 0.16), (0.16, 0.18), (0.18, 0.2), (0.2, 0.22), (0.22, 0.24), (0.24, 0.26), (0.26, 0.28), (0.28, 0.3)]# change
# RMSE_group = [(0.1, 0.12)]
total_molecule_to_bias_RMSE_dict = {}
for RMSE_range in RMSE_group:
    total_molecule_to_bias_RMSE_dict[RMSE_range] = dict()
    RMSE_low = RMSE_range[0]
    RMSE_high = RMSE_range[1]


    # catmap输入energy文件
    energy_filename = "./energies_true.txt"
    # 加了error生成的能量的输出文件夹，注意加上RMSE范围便于识别
    estimate_energy_files_dir = "./output_RMSE_%s_%s_%ssizeN" % (RMSE_low, RMSE_high, energy_num)
    # 尝试创建输出文件夹
    try:
        os.mkdir(estimate_energy_files_dir)
    except:
        pass
    # 生成多少个能量文件

    # ---Step1 读取能量，得到物种名称到能量的字典
    energy_dict = energy_read(energy_filename)
    # print(len(energy_dict))
    # exit(0)
    # 生成的能量最终为字典，key为index，从0到2000，value为另一个字典，species name为key，value为能量
    total_molecule_to_energy_dict = {}

    for i in range(energy_num):
        # print(RMSE_range)
        # print(i)
        # 生成和原始能量相同大小的均匀误差，这些误差的RMSE在一定范围内
        error_info = generate_Random_Number(RMSE_low, RMSE_high, len(energy_dict))
        error = error_info[0]
        actual_bias_RMSE = error_info[1]
        echo = generate_Random_Number(RMSE_low, RMSE_high, len(energy_dict))# 用于跳过一个loop追踪seed，修bug的，
        # print(error)
        # print(generate_Random_Number(RMSE_low, RMSE_high, len(energy_dict)))
        # 然后把新的能量写成字典，按照key排个序，然后按照顺序加上误差
        new_energy = copy.deepcopy(energy_dict)
        new_energy_key = sorted(list(new_energy.keys()))

        # exit(0)
        for j in range(len(new_energy)):
            new_energy[new_energy_key[j]] += error[j]

        total_molecule_to_energy_dict[i] = new_energy
        total_molecule_to_bias_RMSE_dict[RMSE_range][i] = actual_bias_RMSE



    # 基于原来的文本上修改，所以先读取已有数据作为ref data，之后deepcopy
    with open(energy_filename, "r") as f:
        ref_data = f.readlines()
    for i in tqdm.trange(energy_num):
        filename = estimate_energy_files_dir + "\\energies%s.txt" % i
        data = copy.deepcopy(ref_data)
        for index in range(len(data)):
            line = data[index]
            if line.startswith("surface"): continue
            line_data = line.split("\t")
            # 修改元素名称
            line_data[0] = "Ni"
            # 获得当前site name，只要不是gas就可以更改
            if line_data[1] == "gas": continue
            # species名称，用于获得pred结果的能量
            specie_name = line_data[2]
            # 不需要修改H-OH的能量，因为用不到
            if specie_name == "H-OH": continue
            specie_energy = float(total_molecule_to_energy_dict[i][specie_name])
            # line_data[3] = "%.3f" % specie_energy
            line_data[3] = "%.9f" % specie_energy
            # line_data[3] = specie_energy
            # 最后记得修改data
            data[index] = "\t".join(line_data)
        with open(filename, "w") as f:
            f.write("".join(data))

excel_write(total_molecule_to_bias_RMSE_dict, excel_output)
