import pandas as pd
import os
import openpyxl
import datetime
import species

true_energy = './energies_true.txt'
e_path = './2000sizeN/'
# e_path = './entest'
output_path = './'
species = species.species



def energy_read(catmap_input_file):  # 创建一个species对formation energy的dictionary
    with open(catmap_input_file) as f:
        data = f.read().split("\n")
        title = data[0].split("\t")
        species_name = []
        output_data = {}
        for i in data[1:]:
            _data = i.split("\t")
            # FIXME：这里需要排除气态的能量
            if _data[title.index("site_name")] == "gas": continue
            output_data[_data[title.index("species_name")]] = float(_data[title.index("formation_energy")])
    return output_data

def excel_write(all_dic, output_path, species):
    xls = openpyxl.Workbook()
    sheet = xls.create_sheet(index=0, title='error')

    sheet.cell(1, 1).value = 'RMSE Range'
    sheet.cell(1, 2).value = 'order'
    order_one = 3
    for sp in species:
        sheet.cell(1, order_one).value = sp
        order_one += 1

    # all_dic = {(RMSE_low, RMSE_high):{order:{species:energy}}}
    row = 2
    for RMSE in sorted(all_dic.keys()):        
        for order in sorted(all_dic[RMSE].keys()):
            sheet.cell(row, 1).value = str(RMSE)
            sheet.cell(row, 2).value = order
            col = 3
            for spe in all_dic[RMSE][order].keys():
                sheet.cell(row, col).value = all_dic[RMSE][order][spe]
                col += 1
            row += 1


    file_name = output_path + '/test' + '_{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now()) + '.xlsx'
    xls.save(file_name)
    print('excel has been created!')
    exit(0)

#     # all_dic = {'species_name':{'RMSE_range':{divsion_range:count}}}
#     RMSE_order = {}
#     order = 2
#     for key1 in sorted(all_dic.keys()):
#         for key2 in sorted(all_dic[key1].keys()):
#             RMSE_order[order] = key2
#             order += 1

#     xls = openpyxl.Workbook()

#     index_number = 0
#     for i in sorted(all_dic.keys()):
#         sheet = xls.create_sheet(index=index_number, title=i)
#         index_number += 1

#         point_row = 2

#         for point in division_point:
#             sheet.cell(point_row,1).value = point
#             for j in range(len(all_dic[i].keys())):
#                 sheet.cell(1, j + 2).value = 'RMSE ' + str(eval(RMSE_order[j + 2])[0]) + '-' + str(eval(RMSE_order[j + 2])[1]) + 'eV' # 写第一行RMSE范围
#                 sheet.cell(point_row, j + 2).value = all_dic[i][RMSE_order[j + 2]][point]
#             point_row += 1
# else:
#     # all_dic = {'species_name':{'RMSE_range':[error or energy]}}
#     xls = openpyxl.Workbook()
#     index_number = 0
#     for i in sorted(all_dic.keys()):
#         sheet = xls.create_sheet(index=index_number, title=i)
#         index_number += 1
#         column_number = 1
#         for j in sorted(all_dic[i].keys()):
#             sheet.cell(1,column_number).value = 'RMSE ' + str(eval(j)[0]) + '-' + str(eval(j)[1]) + 'eV'# 写第一行RMSE范围
#             row_number = 2
#             for error in all_dic[i][j]:
#                 sheet.cell(row_number, column_number).value = error
#                 row_number += 1
#             column_number += 1



# print(energy_read(true_energy))
# for i in energy_read(true_energy).keys():
#     print(i)
# exit(0)
#   print(energy_read(true_energy)[i])
all_dic = {}
for i in os.listdir(e_path):
    print(i)
    RMSE = i.replace('output_RMSE_','').replace('_2000SizeN','')
    RMSE = RMSE.split('_')
    RMSE = (eval(RMSE[0]), eval(RMSE[1]))
    dic = {}
    for txt in os.listdir(e_path + '/' + i):
        file = e_path + '/' + i + '/' + txt
        order = eval(txt.replace('energies','').replace('.txt',''))
        sp_en = energy_read(file)
        dic[order] = sp_en
    try:
        all_dic[RMSE] = dic
    except KeyError:
        all_dic[RMSE].update(dic)

excel_write(all_dic, output_path, species)